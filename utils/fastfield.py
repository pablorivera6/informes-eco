"""Parse FastField submission Excel exports."""
import re
import openpyxl
from datetime import datetime


def parse_submission(file_bytes) -> dict:
    wb = openpyxl.load_workbook(file_bytes, data_only=True)
    data = {}

    # Root sheet: general info
    ws_root = wb["Root"]
    headers = [c.value for c in ws_root[1]]
    values  = [c.value for c in ws_root[2]]
    root    = dict(zip(headers, values))

    data["fecha_informe"]    = _parse_date(root.get("Fecha del informe"))
    data["cliente"]          = root.get("Cliente", "Ecopetrol")
    data["locacion"]         = root.get("Locacion", "")
    data["profesional_lider"]= root.get("Profesional Lider PCC", "")
    data["contrato"]         = root.get("Contrato/Orden de servicio", "")
    data["objeto_contrato"]  = root.get("Objeto contrato", "")
    data["charla_diaria"]    = root.get("Charla Diaria", "")
    data["evento"]           = root.get("Evento ", "")
    data["hora_inicio"]      = root.get("Hora inicio", "")
    data["hora_fin"]         = root.get("Hora Fin", "")

    # HH — puede venir como número (8) o texto ("7.5 Hrs")
    hh_raw = root.get("Horas hombre") or root.get("Horas Hombre") or ""
    data["horas_hombre"] = _parse_hh(hh_raw)

    # subform_1: Avance por item (narrativa)
    data["avance_items_texto"] = _read_subform(wb, "subform_1", "Avance por item")

    # subform_2: Administrativo
    data["administrativo"] = _read_subform(wb, "subform_2", "Administrativo")

    # subform_3: Avance actividades HSE
    data["avance_hse"] = _read_subform(wb, "subform_3", "Avance actividades HSE")

    # subform_4: Ítems de pago con cantidades (nuevo campo estructurado)
    data["items_fastfield"] = _read_item_quantities(wb)

    # Photos
    data["fotos"] = _read_photos(wb)

    return data


# ── Parsers ───────────────────────────────────────────────────────────────────

def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, str):
        for fmt in ("%m-%d-%Y", "%Y-%m-%d", "%m/%d/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(val.split()[0], fmt).date()
            except ValueError:
                continue
    return None


def _parse_hh(val) -> float:
    """Parse horas hombre — número directo o texto '7.5 Hrs'."""
    if val is None or val == "":
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    try:
        return float(str(val).replace("Hrs", "").replace("hrs", "").strip())
    except ValueError:
        return 0.0


def _read_subform(wb, sheet_name: str, column_name: str) -> str:
    if sheet_name not in wb.sheetnames:
        return ""
    ws = wb[sheet_name]
    headers = [c.value for c in ws[1]]
    lines = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        val = row_dict.get(column_name, "")
        if val and str(val).strip() not in ("", "Xxxxx", "Xxxx", "Xxx"):
            lines.append(str(val).strip())
    return "\n\n".join(lines)


def _read_item_quantities(wb) -> list[dict]:
    """
    Lee subform_4: ítems de pago con cantidad y unidad.

    Soporta el formato con dimensiones separadas (Cantidad #1 / #2 / #3)
    donde la cantidad final = producto de las dimensiones no vacías.
    También es compatible con el formato antiguo de un solo campo 'Cantidad'.
    """
    if "subform_4" not in wb.sheetnames:
        return []

    ws = wb["subform_4"]
    headers = [c.value for c in ws[1]]
    items   = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        raw_item = row_dict.get("Item de pago ") or row_dict.get("Item de pago") or ""
        unidad   = row_dict.get("Unidad", "")

        # Nuevo formato: Cantidad #1, #2, #3
        c1 = _to_float(row_dict.get("Cantidad #1"))
        c2 = _to_float(row_dict.get("Cantidad #2"))
        c3 = _to_float(row_dict.get("Cantidad #3"))

        # Compatibilidad con formato antiguo de campo único
        if c1 is None:
            c1 = _to_float(row_dict.get("Cantidad"))

        if not raw_item or c1 is None:
            continue

        item_num, descripcion = _parse_item_label(str(raw_item))
        if item_num is None:
            continue

        # Multiplicar las dimensiones que vengan llenas
        cantidad_final = c1
        if c2 is not None and c2 != 0:
            cantidad_final *= c2
        if c3 is not None and c3 != 0:
            cantidad_final *= c3

        items.append({
            "item_num":    item_num,
            "descripcion": descripcion,
            "cantidad":    cantidad_final,
            "unidad":      str(unidad or "").strip(),
        })

    return items


def _to_float(val) -> float | None:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _parse_item_label(label: str) -> tuple[int | None, str]:
    """
    Extrae número de ítem y descripción de strings como:
      "008 — CONCRETO CLASE D...  [M3]"
      "023 — PERFORACIÓN VERTICAL...  [M]"
    """
    # Número al inicio (con ceros): "008", "023", "23", "8"
    m = re.match(r"^\s*0*(\d+)\s*[—\-]+\s*(.*)", label)
    if not m:
        return None, label.strip()

    item_num    = int(m.group(1))
    descripcion = m.group(2).strip()

    # Eliminar el [UNIDAD] del final si está en la descripción
    descripcion = re.sub(r"\s*\[[^\]]+\]\s*$", "", descripcion).strip()

    return item_num, descripcion


def _read_photos(wb) -> list[dict]:
    photos = []
    if "multiphoto_picker_1" not in wb.sheetnames:
        return photos
    ws = wb["multiphoto_picker_1"]
    headers = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        filename = row_dict.get("Photo", "")
        comment  = row_dict.get("Comment", "")
        if filename:
            photos.append({"filename": filename, "comment": comment or ""})
    return photos
