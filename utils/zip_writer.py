"""
Write cell values directly into an XLSX ZIP archive.
Preserves 100% of format: charts, images, drawings, VML, styles.
Only modifies the minimum necessary XML inside the worksheet files and sharedStrings.
"""
import io
import re
import zipfile
from datetime import date, datetime
from openpyxl.utils import get_column_letter


# ── Date helpers ──────────────────────────────────────────────────────────────

_EXCEL_EPOCH = date(1899, 12, 30)


def to_excel_serial(d) -> int:
    """Convert a date to Excel serial number."""
    if isinstance(d, datetime):
        d = d.date()
    elif isinstance(d, str):
        d = datetime.strptime(d, "%Y-%m-%d").date()
    return (d - _EXCEL_EPOCH).days


# ── Sheet map ─────────────────────────────────────────────────────────────────

def _get_sheet_map(zf: zipfile.ZipFile) -> dict[str, str]:
    """Return {sheet_name: zip_path} from workbook.xml + workbook.xml.rels."""
    wb_data   = zf.read("xl/workbook.xml").decode("utf-8")
    rels_data = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8")

    # Extract rId → target from rels (filter only worksheet type)
    rid_target = {}
    for m in re.finditer(
        r'<Relationship[^>]+Id="([^"]+)"[^>]+Type="[^"]*worksheet[^"]*"[^>]+Target="([^"]+)"',
        rels_data,
    ):
        rid, target = m.group(1), m.group(2)
        if not target.startswith("xl/"):
            target = "xl/" + target
        rid_target[rid] = target

    # Extract sheet name → rId from workbook.xml
    sheet_map = {}
    r_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    for m in re.finditer(
        r'<sheet\s[^>]*name="([^"]+)"[^>]*r:id="([^"]+)"',
        wb_data,
    ):
        name, rid = m.group(1), m.group(2)
        if rid in rid_target:
            sheet_map[name] = rid_target[rid]

    return sheet_map


# ── Shared strings ────────────────────────────────────────────────────────────

def _load_shared_strings(zf: zipfile.ZipFile) -> tuple[list[str], str]:
    """Return (list_of_strings, raw_xml_bytes)."""
    raw = zf.read("xl/sharedStrings.xml").decode("utf-8")
    strings = []
    for m in re.finditer(r"<si>(.*?)</si>", raw, re.DOTALL):
        # Extract all <t> texts and join (handles rich text / runs)
        texts = re.findall(r"<t(?:[^>]*)>(.*?)</t>", m.group(1), re.DOTALL)
        strings.append("".join(_unescape(t) for t in texts))
    return strings, raw


def _escape(text: str) -> str:
    return (
        text.replace("&", "&amp;")
            .replace("<", "&lt;")
            .replace(">", "&gt;")
            .replace('"', "&quot;")
    )


def _unescape(text: str) -> str:
    return (
        text.replace("&amp;", "&")
            .replace("&lt;", "<")
            .replace("&gt;", ">")
            .replace("&quot;", '"')
    )


def _add_shared_string(strings: list[str], ss_xml: str, text: str) -> tuple[int, str]:
    """Add text to sharedStrings if not present; return (index, updated_xml)."""
    if text in strings:
        return strings.index(text), ss_xml

    idx = len(strings)
    strings.append(text)

    # Build new <si> element preserving newlines with xml:space="preserve"
    escaped = _escape(text)
    if "\n" in text:
        new_si = f'<si><t xml:space="preserve">{escaped}</t></si>'
    else:
        new_si = f"<si><t>{escaped}</t></si>"

    # Insert before </sst>
    ss_xml = ss_xml.replace("</sst>", new_si + "</sst>")

    # Update count attribute in <sst ... count="N">
    def inc_count(m):
        return m.group(0).replace(
            f'count="{m.group(1)}"',
            f'count="{int(m.group(1))+1}"',
        )
    ss_xml = re.sub(r'count="(\d+)"', inc_count, ss_xml, count=1)
    ss_xml = re.sub(r'uniqueCount="(\d+)"', lambda m: m.group(0).replace(
        m.group(1), str(int(m.group(1)) + 1)
    ), ss_xml, count=1)

    return idx, ss_xml


# ── Cell update engine ────────────────────────────────────────────────────────

def _cell_ref(col: int, row: int) -> str:
    return get_column_letter(col) + str(row)


def _val_str(value: float | int) -> str:
    if isinstance(value, int) or (isinstance(value, float) and value == int(value)):
        return str(int(value))
    return repr(value)


# Matches both <c r="REF" ...>...</c> and <c r="REF" .../>
_CELL_RE_TEMPLATE = (
    r'<c r="{ref}"([^/]*?)(?:/>|>(.*?)</c>)'
)


def _get_cell_style(sheet_xml: str, ref: str) -> str:
    """Return the style attribute (e.g. ' s="356"') of a cell, or '' if none."""
    result = _find_cell(sheet_xml, ref)
    if not result:
        return ""
    attrs = result[1]
    m = re.search(r'\ss="\d+"', attrs)
    return m.group(0) if m else ""


def _find_cell(sheet_xml: str, ref: str):
    """Return (match, open_attrs, content, is_self_closing) or None."""
    pat = re.compile(
        r'<c r="' + re.escape(ref) + r'"([^/]*?)(?:/>|>(.*?)</c>)',
        re.DOTALL,
    )
    m = pat.search(sheet_xml)
    if not m:
        return None
    attrs = m.group(1)            # everything between ref and /> or >
    content = m.group(2) or ""   # None if self-closing
    is_self_closing = m.group(2) is None
    return m, attrs, content, is_self_closing


def _set_cell_number(sheet_xml: str, ref: str, value: float | int) -> str:
    """Set numeric value in a cell. Handles self-closing empty cells."""
    val_s = _val_str(value)
    result = _find_cell(sheet_xml, ref)
    if result:
        m, attrs, content, is_sc = result
        # Remove t="s" from attrs if present
        attrs = re.sub(r'\s*t="s"', "", attrs)
        if "<f>" in content:
            # Keep formula, update cached value
            new_content = re.sub(r"<v>[^<]*</v>", f"<v>{val_s}</v>", content)
            if "<v>" not in new_content:
                new_content += f"<v>{val_s}</v>"
        else:
            new_content = f"<v>{val_s}</v>"
        replacement = f'<c r="{ref}"{attrs}>{new_content}</c>'
        sheet_xml = sheet_xml[:m.start()] + replacement + sheet_xml[m.end():]
    else:
        sheet_xml = _insert_cell(sheet_xml, ref, f"<v>{val_s}</v>", type_attr="")
    return sheet_xml


def _set_cell_string(sheet_xml: str, ref: str, ss_idx: int) -> str:
    """Set a shared-string value in a cell (type t="s")."""
    result = _find_cell(sheet_xml, ref)
    if result:
        m, attrs, content, is_sc = result
        # Ensure t="s"
        if 't="s"' not in attrs:
            attrs = re.sub(r't="[^"]*"', "", attrs)
            attrs = attrs.rstrip() + ' t="s"'
        replacement = f'<c r="{ref}"{attrs}><v>{ss_idx}</v></c>'
        sheet_xml = sheet_xml[:m.start()] + replacement + sheet_xml[m.end():]
    else:
        sheet_xml = _insert_cell(
            sheet_xml, ref, f"<v>{ss_idx}</v>", type_attr=' t="s"'
        )
    return sheet_xml


def _insert_cell(sheet_xml: str, ref: str, value_xml: str, type_attr: str = "") -> str:
    """Insert a new <c> element into the correct <row>, creating the row if needed."""
    col_str = "".join(c for c in ref if c.isalpha())
    row_num = int("".join(c for c in ref if c.isdigit()))
    from openpyxl.utils import column_index_from_string
    col_num = column_index_from_string(col_str)

    row_pattern = re.compile(
        r'(<row r="' + str(row_num) + r'"[^>]*>)(.*?)(</row>)',
        re.DOTALL,
    )
    m = row_pattern.search(sheet_xml)
    if not m:
        sheet_xml = _create_row(sheet_xml, row_num)
        m = row_pattern.search(sheet_xml)
        if not m:
            return sheet_xml

    open_tag, row_content, close_tag = m.group(1), m.group(2), m.group(3)
    new_cell = f'<c r="{ref}"{type_attr}>{value_xml}</c>'

    # Insert in correct column order
    existing_cells = list(re.finditer(r'<c r="([A-Z]+)(\d+)"', row_content))
    insert_pos = len(row_content)
    for cell_m in existing_cells:
        from openpyxl.utils import column_index_from_string as c2i
        if c2i(cell_m.group(1)) > col_num:
            insert_pos = cell_m.start()
            break

    row_content = row_content[:insert_pos] + new_cell + row_content[insert_pos:]
    sheet_xml = sheet_xml[:m.start()] + open_tag + row_content + close_tag + sheet_xml[m.end():]
    return sheet_xml


def _create_row(sheet_xml: str, row_num: int) -> str:
    """Create an empty <row> element in the correct position within <sheetData>."""
    sd_m = re.search(r"(</sheetData>)", sheet_xml)
    if not sd_m:
        return sheet_xml

    insert_pos = sd_m.start()

    row_positions = list(re.finditer(r'<row r="(\d+)"', sheet_xml[:sd_m.start()]))
    for rm in reversed(row_positions):
        if int(rm.group(1)) < row_num:
            close_m = re.search(r'</row>', sheet_xml[rm.end():sd_m.start()])
            if close_m:
                insert_pos = rm.end() + close_m.end()
            break

    new_row = f'<row r="{row_num}">'
    sheet_xml = sheet_xml[:insert_pos] + new_row + "</row>" + sheet_xml[insert_pos:]
    return sheet_xml


# ── High-level writer ─────────────────────────────────────────────────────────

class XlsxZipWriter:
    """
    Modify specific cells in an XLSX without touching any other content.
    Usage:
        writer = XlsxZipWriter(xlsx_bytes)
        writer.set_number("Resumen", 10, 14, 136)        # Reporte No.
        writer.set_date("Resumen", 10, 19, date(2026,5,8))
        writer.set_text("Resumen", 18, 6, "Avance...")
        result_bytes = writer.save()
    """

    def __init__(self, xlsx_bytes: bytes):
        if hasattr(xlsx_bytes, "read"):
            xlsx_bytes = xlsx_bytes.read()
        self._original = xlsx_bytes
        with zipfile.ZipFile(io.BytesIO(xlsx_bytes), "r") as zf:
            self._sheet_map = _get_sheet_map(zf)
            self._shared_strings, self._ss_xml = _load_shared_strings(zf)

        # path → modified xml string (start as None = unchanged)
        self._modified_sheets: dict[str, str] = {}
        # path → modified binary (media files)
        self._modified_media: dict[str, bytes] = {}

    def _get_sheet_xml(self, sheet_name: str) -> str:
        path = self._sheet_map[sheet_name]
        if path not in self._modified_sheets:
            with zipfile.ZipFile(io.BytesIO(self._original), "r") as zf:
                self._modified_sheets[path] = zf.read(path).decode("utf-8")
        return self._modified_sheets[path]

    def _save_sheet_xml(self, sheet_name: str, xml: str):
        path = self._sheet_map[sheet_name]
        self._modified_sheets[path] = xml

    def set_number(self, sheet: str, row: int, col: int, value: float | int):
        xml = self._get_sheet_xml(sheet)
        ref = _cell_ref(col, row)
        xml = _set_cell_number(xml, ref, value)
        self._save_sheet_xml(sheet, xml)

    def set_date(self, sheet: str, row: int, col: int, value):
        serial = to_excel_serial(value)
        self.set_number(sheet, row, col, serial)

    def set_text(self, sheet: str, row: int, col: int, text: str):
        if not text:
            return
        idx, self._ss_xml = _add_shared_string(self._shared_strings, self._ss_xml, text)
        xml = self._get_sheet_xml(sheet)
        ref = _cell_ref(col, row)
        xml = _set_cell_string(xml, ref, idx)
        self._save_sheet_xml(sheet, xml)

    def find_date_col(self, sheet: str, target_date, date_row: int = 1) -> int | None:
        """Find the column index where date_row contains target_date."""
        if isinstance(target_date, datetime):
            target_date = target_date.date()
        serial = to_excel_serial(target_date)
        xml = self._get_sheet_xml(sheet)

        row_m = re.search(
            r'<row r="' + str(date_row) + r'"[^>]*>(.*?)</row>', xml, re.DOTALL
        )
        if not row_m:
            return None
        row_xml = row_m.group(1)

        # Match each cell's OWN <v> bounded by its </c>; an empty self-closing
        # cell must not borrow the next cell's value (that returns a wrong col).
        pat = re.compile(
            r'<c r="([A-Z]+)' + str(date_row) + r'"[^>]*?(?:/>|>(.*?)</c>)',
            re.DOTALL,
        )
        for m in pat.finditer(row_xml):
            content = m.group(2)
            if not content:
                continue
            v = re.search(r"<v>(\d+)</v>", content)
            if v and int(v.group(1)) == serial:
                from openpyxl.utils import column_index_from_string
                return column_index_from_string(m.group(1))
        return None

    def clone_row_format(self, sheet: str, src_row: int) -> int:
        """
        Clone an existing row's structure (every <c> with its style) into a brand
        new, empty row appended at the end of the sheet. Values and formulas are
        stripped so the caller can fill it; cell styles (borders, date/number
        formats) are preserved so the new day row looks like the rest of the
        calendar. Returns the new row number.
        """
        xml = self._get_sheet_xml(sheet)
        src_m = re.search(
            r'<row r="' + str(src_row) + r'"[^>]*>(.*?)</row>', xml, re.DOTALL
        )
        if not src_m:
            return src_row  # nothing to clone

        new_row = max(int(r) for r in re.findall(r'<row r="(\d+)"', xml)) + 1

        # Rebuild each source cell at the new row, keeping only its style attr.
        cells = []
        for cm in re.finditer(
            r'<c r="([A-Z]+)\d+"([^>]*?)(?:/>|>.*?</c>)', src_m.group(1), re.DOTALL
        ):
            col_ltr, attrs = cm.group(1), cm.group(2)
            s = re.search(r'\ss="\d+"', attrs)
            style = s.group(0) if s else ""
            cells.append(f'<c r="{col_ltr}{new_row}"{style}/>')

        new_row_xml = f'<row r="{new_row}">' + "".join(cells) + "</row>"
        xml = re.sub(r"(</sheetData>)", new_row_xml + r"\1", xml, count=1)
        self._save_sheet_xml(sheet, xml)
        return new_row

    def get_number(self, sheet: str, row: int, col: int) -> float | None:
        """Read the cached numeric value of a cell (None if missing or non-numeric)."""
        xml = self._get_sheet_xml(sheet)
        ref = _cell_ref(col, row)
        result = _find_cell(xml, ref)
        if not result:
            return None
        v_m = re.search(r"<v>([^<]*)</v>", result[2])
        if not v_m or not v_m.group(1):
            return None
        try:
            return float(v_m.group(1))
        except ValueError:
            return None

    def add_to_number(self, sheet: str, row: int, col: int, delta: float):
        """Add delta to existing numeric cell value (for C.Control daily quantities)."""
        xml = self._get_sheet_xml(sheet)
        ref = _cell_ref(col, row)
        result = _find_cell(xml, ref)
        if result:
            _, attrs, content, is_sc = result
            v_match = re.search(r"<v>([^<]*)</v>", content)
            current = float(v_match.group(1)) if v_match and v_match.group(1) else 0.0
        else:
            current = 0.0
        xml = _set_cell_number(xml, ref, current + delta)
        self._save_sheet_xml(sheet, xml)

    def extend_curva_s(self, sheet: str, target_date_col: int):
        """
        Extiende las filas 8 (Avance Acumulado) y 9 (Avance diario) en C.Control
        desde la última columna con fórmula hasta target_date_col, inclusive.

        Fila 9: SUMPRODUCT({col}15:{col}128,$G$15:$G$128)/$I$13  [s=347]
        Fila 8: {col}9+{prev_col}8                               [s=356]
        """
        from openpyxl.utils import get_column_letter, column_index_from_string

        xml = self._get_sheet_xml(sheet)

        # Encontrar la última columna con fórmula en fila 9
        row9_m = re.search(r'<row r="9"[^>]*>(.*?)</row>', xml, re.DOTALL)
        if not row9_m:
            return

        # Todas las referencias de columna que tienen contenido en fila 9
        cols_in_row9 = [
            column_index_from_string(m.group(1))
            for m in re.finditer(r'<c r="([A-Z]+)9"', row9_m.group(1))
        ]
        last_col = max(cols_in_row9) if cols_in_row9 else 16  # P = col 16 mínimo

        if last_col >= target_date_col:
            return  # Ya llega hasta la fecha o más allá

        # Agregar fórmulas para cada columna faltante
        for col in range(last_col + 1, target_date_col + 1):
            col_ltr  = get_column_letter(col)
            prev_ltr = get_column_letter(col - 1)

            # Estilo (formato %) tomado de la columna previa para conservar el
            # formato de porcentaje en vez de quedar en formato "General".
            style9 = _get_cell_style(xml, f"{prev_ltr}9") or ' s="347"'
            style8 = _get_cell_style(xml, f"{prev_ltr}8") or ' s="356"'

            # Fila 9: avance diario
            f9  = f"SUMPRODUCT({col_ltr}15:{col_ltr}128,$G$15:$G$128)/$I$13"
            ref9 = f"{col_ltr}9"
            result9 = _find_cell(xml, ref9)
            if result9:
                m9, _, _, _ = result9
                replacement9 = f'<c r="{ref9}"{style9}><f>{f9}</f></c>'
                xml = xml[:m9.start()] + replacement9 + xml[m9.end():]
            else:
                xml = _insert_cell(xml, ref9, f"<f>{f9}</f>", type_attr=style9)

            # Fila 8: avance acumulado
            f8  = f"{col_ltr}9+{prev_ltr}8"
            ref8 = f"{col_ltr}8"
            result8 = _find_cell(xml, ref8)
            if result8:
                m8, _, _, _ = result8
                replacement8 = f'<c r="{ref8}"{style8}><f>{f8}</f></c>'
                xml = xml[:m8.start()] + replacement8 + xml[m8.end():]
            else:
                xml = _insert_cell(xml, ref8, f"<f>{f8}</f>", type_attr=style8)

        self._save_sheet_xml(sheet, xml)

    def accumulate_recursos_hh(self,
                               sheet: str = "Recursos",
                               hh_col: int = 5,    # columna E  → E$11 (Cantidad Horas día)
                               data_col: int = 6,   # columna F  → CUSIANA acumulado
                               hh_row: int = 11,
                               row_start: int = 15,
                               row_end: int = 46):
        """
        Simula copiar la columna fórmula (=F_row+E$11) y pegarla en F.
        Para cada fila con valor > 0 en columna F, añade E11 (horas del día).
        """
        if sheet not in self._sheet_map:
            return
        xml = self._get_sheet_xml(sheet)

        # Leer E11 (Cantidad Horas día)
        e11_ref = _cell_ref(hh_col, hh_row)
        res = _find_cell(xml, e11_ref)
        if not res:
            return
        v_m = re.search(r"<v>([^<]*)</v>", res[2])
        if not v_m or not v_m.group(1):
            return
        try:
            daily_hh = float(v_m.group(1))
        except ValueError:
            return
        if daily_hh <= 0:
            return

        # Sumar daily_hh a cada fila no vacía de columna F
        for row in range(row_start, row_end + 1):
            f_ref = _cell_ref(data_col, row)
            res = _find_cell(xml, f_ref)
            if not res:
                continue
            v_m = re.search(r"<v>([^<]*)</v>", res[2])
            if not v_m or not v_m.group(1):
                continue
            try:
                current = float(v_m.group(1))
            except ValueError:
                continue
            if current <= 0:
                continue
            xml = _set_cell_number(xml, f_ref, current + daily_hh)

        self._save_sheet_xml(sheet, xml)

    def replace_photo(self, slot: int, image_bytes: bytes):
        """
        Reemplaza una foto en el registro fotográfico de Resumen.
        slot: 1-6 → xl/media/image4.png a image9.png
        Convierte cualquier formato a PNG para compatibilidad.
        """
        from PIL import Image as PILImage
        img = PILImage.open(io.BytesIO(image_bytes))
        buf = io.BytesIO()
        img.convert("RGB").save(buf, format="PNG")
        self._modified_media[f"xl/media/image{slot + 3}.png"] = buf.getvalue()

    def blank_photo(self, slot: int):
        """Deja en blanco el placeholder de una foto (slot 1-6) escribiendo una
        imagen blanca, para que no arrastre la foto del reporte anterior."""
        from PIL import Image as PILImage
        buf = io.BytesIO()
        PILImage.new("RGB", (800, 600), "white").save(buf, format="PNG")
        self._modified_media[f"xl/media/image{slot + 3}.png"] = buf.getvalue()

    def save(self) -> bytes:
        output = io.BytesIO()
        with zipfile.ZipFile(io.BytesIO(self._original), "r") as zin:
            with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    if item.filename == "xl/sharedStrings.xml":
                        data = self._ss_xml.encode("utf-8")
                    elif item.filename in self._modified_sheets:
                        data = self._modified_sheets[item.filename].encode("utf-8")
                    elif item.filename in self._modified_media:
                        data = self._modified_media[item.filename]
                    else:
                        data = zin.read(item.filename)
                    zout.writestr(item, data)
        output.seek(0)
        return output.read()
