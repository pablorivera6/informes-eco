"""Read and write operations on the formal Ecopetrol report Excel."""
import copy
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime, date


# ── C.Control parsing ────────────────────────────────────────────────────────

def read_c_control_items(wb) -> list[dict]:
    """Return all contract items from C.Control with their row numbers."""
    ws = wb["C.Control"]
    items = []
    section = "CUSIANA"  # First section (labeled PROYECTO in Excel) = Cusiana

    for row_num in range(11, ws.max_row + 1):
        col_b = ws.cell(row=row_num, column=2).value

        # Section boundaries detected by col B value
        if col_b in ("FLORENA", "FLOREÑA"):
            section = "FLORENA"
            continue
        if col_b == "CUPIAGUA":
            section = "CUPIAGUA"
            continue

        # Item rows have an integer in col B (item number)
        if not isinstance(col_b, (int, float)) or col_b != int(col_b):
            continue
        item_num = int(col_b)

        especialidad = ws.cell(row=row_num, column=3).value
        codigo_sap   = ws.cell(row=row_num, column=4).value
        descripcion  = ws.cell(row=row_num, column=5).value
        unidad       = ws.cell(row=row_num, column=6).value
        valor_unit   = ws.cell(row=row_num, column=7).value
        cantidad_tot = ws.cell(row=row_num, column=8).value

        # ACUMULADO is col L (12) — SUM formula; read cached value
        acum_cell = ws.cell(row=row_num, column=12)
        acumulado = None
        if not str(acum_cell.value or "").startswith("="):
            acumulado = acum_cell.value

        items.append({
            "section": section,
            "item": item_num,
            "especialidad": especialidad or "",
            "codigo_sap": str(codigo_sap or ""),
            "descripcion": descripcion or "",
            "unidad": unidad or "",
            "valor_unitario": valor_unit,
            "cantidad_total": cantidad_tot,
            "acumulado": acumulado,
            "row_num": row_num,
        })

    return items


def find_date_column(wb, target_date) -> int | None:
    """Return the column index in C.Control for a given date (row 1 header)."""
    ws = wb["C.Control"]
    if isinstance(target_date, datetime):
        target_date = target_date.date()
    elif isinstance(target_date, str):
        target_date = datetime.strptime(target_date, "%Y-%m-%d").date()

    for col in range(16, ws.max_column + 1):
        val = ws.cell(row=1, column=col).value
        if val is None:
            continue
        if isinstance(val, datetime) and val.date() == target_date:
            return col
        if isinstance(val, date) and val == target_date:
            return col
    return None


# ── Report writing ────────────────────────────────────────────────────────────

# Celdas de descripción fotográfica en Resumen: (fila, col) por slot 1-6
_PHOTO_DESC_CELLS = [
    (32, 2),   # Foto 1 — izquierda
    (32, 17),  # Foto 2 — derecha
    (34, 2),   # Foto 3 — izquierda
    (34, 17),  # Foto 4 — derecha
    (36, 2),   # Foto 5 — izquierda
    (36, 17),  # Foto 6 — derecha
]


def update_report(template_bytes, form_data: dict, item_quantities: list[dict]) -> bytes:
    """
    Write cell values into the XLSX using direct ZIP manipulation.
    Preserves 100% of original format: charts, images, drawings, VML, styles.
    """
    from utils.zip_writer import XlsxZipWriter
    from io import BytesIO

    if hasattr(template_bytes, "read"):
        template_bytes = template_bytes.read()

    w = XlsxZipWriter(template_bytes)
    target_date = form_data.get("fecha_informe")

    # ── Resumen ──────────────────────────────────────────────────────────────
    if form_data.get("reporte_no") is not None:
        w.set_number("Resumen", 10, 14, form_data["reporte_no"])        # N10
    if target_date:
        w.set_date("Resumen", 10, 19, target_date)                      # S10
    if form_data.get("charla_diaria"):
        w.set_text("Resumen", 14, 27, form_data["charla_diaria"])       # AA14
    if form_data.get("avance_cusiana"):
        w.set_text("Resumen", 18, 6, form_data["avance_cusiana"])       # F18
    if form_data.get("administrativo_cusiana"):
        w.set_text("Resumen", 19, 6, form_data["administrativo_cusiana"])  # F19
    if form_data.get("avance_hse_cusiana"):
        w.set_text("Resumen", 20, 6, form_data["avance_hse_cusiana"])   # F20

    _set_if("Resumen", 27,  8, "personal_calificado_region",        w, form_data)  # H27
    _set_if("Resumen", 27, 10, "personal_calificado_no_region",     w, form_data)  # J27
    _set_if("Resumen", 28,  8, "personal_no_calificado_region",     w, form_data)  # H28
    _set_if("Resumen", 28, 10, "personal_no_calificado_no_region",  w, form_data)  # J28
    _set_if("Resumen", 27, 25, "maquinaria",                        w, form_data)  # Y27
    _set_if("Resumen", 28, 25, "equipo",                            w, form_data)  # Y28

    # ── HSE ──────────────────────────────────────────────────────────────────
    if target_date:
        w.set_date("HSE", 10, 7, target_date)                          # G10
    if form_data.get("hh_dia") is not None:
        w.set_number("HSE", 22, 8, form_data["hh_dia"])                # H22
    if form_data.get("charla_diaria"):
        w.set_text("HSE", 24, 8, form_data["charla_diaria"])           # H24

    # HSE daily log table (find row for target_date)
    if target_date:
        hse_log_row = _find_hse_log_row(w, target_date)
        _hse_log_map = {
            "hse_accid_cpt":       3,
            "hse_accid_spt":       4,
            "hse_primeros_aux":    5,
            "hse_derrames":        6,
            "hse_incid_viales":    7,
            "hse_casi_accid":      8,
            "hse_fallas_ctrl":     9,
            "hse_aseguramiento":  10,
            "hse_visitas_ger":    11,
            "hse_alcoholimetrias":12,
            "hh_dia":             14,
        }
        for key, col in _hse_log_map.items():
            val = form_data.get(key)
            if val is not None:
                w.set_number("HSE", hse_log_row, col, int(val) if isinstance(val, float) and val == int(val) else val)
        if form_data.get("charla_diaria"):
            w.set_text("HSE", hse_log_row, 15, form_data["charla_diaria"])

    # ── C.Control ────────────────────────────────────────────────────────────
    if target_date:
        date_col = w.find_date_col("C.Control", target_date)
        if date_col:
            # Escribir cantidades de ítems
            for iq in (item_quantities or []):
                qty = iq.get("cantidad_final", 0) or 0
                if qty > 0:
                    w.add_to_number("C.Control", iq["row_num"], date_col, qty)

            # Extender filas 8 y 9 (Curva S) hasta la fecha del informe
            w.extend_curva_s("C.Control", date_col)

            # Avance acumulado: escribir en C.Control, Resumen y Curvas
            avance_acum = form_data.get("avance_real_acumulado")

            # Cached value en C.Control fila 8 (preserva la fórmula)
            if avance_acum is not None:
                w.set_number("C.Control", 8, date_col, avance_acum)

            if "Curvas" in w._sheet_map:
                # Fechas de Curvas están en fila 1
                curvas_col = w.find_date_col("Curvas", target_date, date_row=1)
                if curvas_col:
                    # Real acumulado → Curvas fila 4: escribir la FÓRMULA que
                    # referencia C.Control (no el número), igual que el resto de
                    # la fila ('C.Control'!{col}8). El valor cacheado permite que
                    # se vea el % antes de que Excel recalcule.
                    if avance_acum is not None:
                        from openpyxl.utils import get_column_letter
                        cc_col = get_column_letter(date_col)
                        w.set_formula(
                            "Curvas", 4, curvas_col,
                            f"'C.Control'!{cc_col}8", cached=avance_acum,
                        )
                    # Plan acumulado → leer de Curvas fila 3 → Resumen B14
                    plan_acum = w.get_number("Curvas", 3, curvas_col)
                    if plan_acum is not None:
                        w.set_number("Resumen", 14, 2, plan_acum)

            # Resumen E14 = avance real acumulado
            if avance_acum is not None:
                w.set_number("Resumen", 14, 5, avance_acum)

    # ── Fotos ──────────────────────────────────────────────────────────────────
    _update_fotos(w, form_data)

    # ── Recursos — acumular HH del día en columna F ─────────────────────────
    w.accumulate_recursos_hh()

    return w.save()


def _update_fotos(w, form_data: dict):
    """Escribe descripciones y reemplaza imágenes del registro fotográfico."""
    fotos = form_data.get("fotos", [])  # lista de {descripcion, image_bytes}
    fecha_str = ""
    if form_data.get("fecha_informe"):
        d = form_data["fecha_informe"]
        fecha_str = f"{d.day:02d}/{d.month:02d}/{d.year}"
    locacion = form_data.get("locacion_display", "Cusiana")

    for idx, (row, col) in enumerate(_PHOTO_DESC_CELLS):
        foto = fotos[idx] if idx < len(fotos) else {}
        img_bytes = foto.get("image_bytes")

        if img_bytes:
            w.replace_photo(idx + 1, img_bytes)  # slots 1-6
            desc = foto.get("descripcion", "").strip()
            texto = f"Fecha: {fecha_str}\nUbicación: {locacion}\nDescripción: {desc}"
            w.set_text("Resumen", row, col, texto)
        else:
            # Sin foto en este espacio: dejar en blanco (imagen y descripción)
            # para que no arrastre la foto/descr. del reporte anterior.
            w.blank_photo(idx + 1)
            w.set_text("Resumen", row, col, " ")


def _set_if(sheet, row, col, key, writer, form_data):
    val = form_data.get(key)
    if val is not None:
        writer.set_number(sheet, row, col, val)


def _find_hse_log_row(writer, target_date) -> int:
    """Find HSE log row by searching column B from row 30 for target_date serial."""
    from utils.zip_writer import to_excel_serial
    serial = to_excel_serial(target_date)

    xml = writer._get_sheet_xml("HSE")
    import re
    # Search for existing date in col B (rows 30+). Match each cell's OWN <v>
    # bounded by its </c>; a value-less self-closing cell (e.g. weekend spacer
    # rows) must NOT borrow the next cell's value, or we'd return a wrong row.
    for m in re.finditer(
        r'<c r="B(\d+)"[^>]*?(?:/>|>(.*?)</c>)', xml, re.DOTALL
    ):
        row_num = int(m.group(1))
        if row_num < 30:
            continue
        content = m.group(2)
        if not content:
            continue
        v = re.search(r"<v>(\d+)</v>", content)
        if v and int(v.group(1)) == serial:
            return row_num

    # Date not present yet (e.g. a new day past the last filled row).
    # The calendar pre-builds empty day rows (B cell exists but has no value),
    # interleaved with spacer/week-header rows (column A holds the week number).
    # We must insert chronologically: scan forward from the row right after the
    # latest date that is < target, and take the first EMPTY DAY row — i.e. a row
    # whose B cell has no value AND whose A cell has no value (spacers carry a
    # week number in A, so skip them).
    def _cell_value(col: str, row: int):
        m = re.search(
            r'<c r="' + col + str(row) + r'"[^>]*?(?:/>|>(.*?)</c>)',
            xml, re.DOTALL,
        )
        if not m:
            return None, False  # cell absent
        content = m.group(1)
        if not content:
            return "", True  # cell present but empty
        v = re.search(r"<v>(\d+)</v>", content)
        return (v.group(1) if v else ""), True

    # Row holding the closest date below the target. Pick by largest DATE (not
    # by row number) so the monthly-summary block at the bottom — which repeats
    # earlier dates at high row numbers — never wins over the real predecessor.
    best_serial, best_row = -1, None
    for m in re.finditer(r'<c r="B(\d+)"[^>]*?(?:/>|>(.*?)</c>)', xml, re.DOTALL):
        rn = int(m.group(1))
        if rn < 30 or not m.group(2):
            continue
        v = re.search(r"<v>(\d+)</v>", m.group(2))
        if v:
            s = int(v.group(1))
            if s < serial and s > best_serial:
                best_serial, best_row = s, rn
    start = (best_row + 1) if best_row else 30

    for r in range(start, start + 120):
        b_val, b_exists = _cell_value("B", r)
        if not b_exists:
            break  # reached the end of the contiguous table
        if b_val:
            break  # hit a dated row (monthly-summary block) → no empty days left
        a_val, _ = _cell_value("A", r)
        if a_val:
            continue  # spacer / week-header row
        writer.set_date("HSE", r, 2, target_date)  # B col = col 2
        return r

    # No pre-built empty day row available: clone the predecessor day row's
    # formatting into a fresh row appended at the end, then date it. This keeps
    # the report working indefinitely once the pre-built calendar is exhausted.
    template_row = best_row if best_row else 30
    new_row = writer.clone_row_format("HSE", template_row)
    writer.set_date("HSE", new_row, 2, target_date)
    return new_row


def read_reporte_no(wb) -> int:
    """Read current Reporte No. from Resumen."""
    ws = wb["Resumen"]
    val = ws.cell(row=10, column=14).value
    try:
        return int(val)
    except (TypeError, ValueError):
        return 0
