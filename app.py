"""Automatización de Informes Diarios — Proyecto Ecopetrol CW309754."""
import base64
import io
from datetime import date
from pathlib import Path

import openpyxl
import pandas as pd
import streamlit as st

from utils.fastfield import parse_submission
from utils.fastfield_api import download_submission_photos
from utils.excel_ops import (
    find_date_column,
    read_c_control_items,
    read_reporte_no,
    update_report,
)


# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Informes Diarios — Ecopetrol CW309754",
    layout="wide",
    initial_sidebar_state="collapsed",
)


# ── Location mapping ──────────────────────────────────────────────────────────
LOCATION_MAP: dict[str, str] = {
    "cusiana":   "CUSIANA",
    "florena":   "FLORENA",
    "floreña":   "FLORENA",
    "cupiagua":  "CUPIAGUA",
    "chupiguay": "CUPIAGUA",
}
SECTION_LABELS = {
    "CUSIANA":  "Cusiana",
    "FLORENA":  "Floreña",
    "CUPIAGUA": "Cupiagua / Chupiguay",
}

def detect_section(locacion: str) -> str | None:
    return LOCATION_MAP.get((locacion or "").strip().lower())

# ── Asset helpers ─────────────────────────────────────────────────────────────
def _img_b64(path: str) -> str:
    data = Path(path).read_bytes()
    ext  = path.rsplit(".", 1)[-1].lower()
    mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
    return f"data:{mime};base64,{base64.b64encode(data).decode()}"

LOGO_PCC = _img_b64("assets/logo_pcc.png")
LOGO_ECO = _img_b64("assets/logo_ecopetrol.png")

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Lexend:wght@400;500;600;700&family=Inter:wght@400;500;600&display=swap');

/* ─── Reset & base ─── */
*, *::before, *::after { box-sizing: border-box; }

.stApp {
    background-color: #0D1117 !important;
    color: #E6EDF3 !important;
    font-family: 'Inter', -apple-system, BlinkMacSystemFont, 'Segoe UI', sans-serif !important;
    font-size: 14px !important;
    line-height: 1.5 !important;
}

/* Hide Streamlit chrome */
[data-testid="stHeader"],
#MainMenu,
footer,
[data-testid="stToolbar"],
[data-testid="stDecoration"] {
    display: none !important;
}

/* Content container */
.main .block-container {
    padding: 0 2.5rem 5rem !important;
    max-width: 1160px !important;
}

/* ─── Header ─── */
.pcc-header {
    display: flex;
    align-items: center;
    justify-content: space-between;
    background: #161B22;
    border-bottom: 2px solid #BE1E2D;
    padding: 18px 2.5rem;
    margin: 0 -2.5rem 2.5rem;
}
.pcc-header .h-logos {
    display: flex;
    align-items: center;
    gap: 0;
}
.pcc-header .h-title {
    text-align: center;
    flex: 1;
    padding: 0 2rem;
}
.pcc-header .h-title h1 {
    margin: 0;
    font-family: 'Lexend', sans-serif;
    font-size: 18px;
    font-weight: 600;
    color: #E6EDF3;
    letter-spacing: -0.3px;
    line-height: 1.2;
}
.pcc-header .h-title p {
    margin: 4px 0 0;
    font-size: 11px;
    color: #6E7681;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    font-weight: 500;
}

/* ─── Step headers ─── */
.step-hdr {
    display: flex;
    align-items: center;
    gap: 16px;
    margin: 2.5rem 0 1.25rem;
    padding-bottom: 14px;
    border-bottom: 1px solid #21262D;
}
.step-num {
    font-family: 'Lexend', sans-serif;
    font-size: 13px;
    font-weight: 700;
    color: #BE1E2D;
    background: #1C1012;
    border: 1px solid #3D1519;
    border-radius: 6px;
    padding: 4px 10px;
    letter-spacing: 0.5px;
    white-space: nowrap;
    flex-shrink: 0;
}
.step-title {
    font-family: 'Lexend', sans-serif;
    font-size: 16px;
    font-weight: 600;
    color: #E6EDF3;
    letter-spacing: -0.2px;
}
.step-desc {
    font-size: 12px;
    color: #6E7681;
    margin-top: 1px;
}

/* ─── Upload zone labels ─── */
.upload-label {
    font-family: 'Lexend', sans-serif;
    font-size: 13px;
    font-weight: 600;
    color: #C9D1D9;
    margin-bottom: 4px;
}
.upload-desc {
    font-size: 12px;
    color: #6E7681;
    margin-bottom: 10px;
}

/* ─── Status pills ─── */
.pill {
    display: inline-flex;
    align-items: center;
    gap: 7px;
    padding: 8px 14px;
    border-radius: 6px;
    font-size: 13px;
    line-height: 1.4;
    width: 100%;
    margin: 6px 0;
}
.pill-ok   { background: #0D1F12; border: 1px solid #1F4A27; color: #56D364; }
.pill-info { background: #0C1E36; border: 1px solid #1A3F6A; color: #58A6FF; }
.pill-warn { background: #1F1A0A; border: 1px solid #4A3A10; color: #D29922; }
.pill-err  { background: #1F0D0D; border: 1px solid #4A1A1A; color: #F85149; }
.pill-dot {
    width: 6px; height: 6px; border-radius: 50%;
    flex-shrink: 0;
}
.pill-ok  .pill-dot { background: #56D364; }
.pill-info .pill-dot { background: #58A6FF; }
.pill-warn .pill-dot { background: #D29922; }
.pill-err  .pill-dot { background: #F85149; }

/* ─── Section sub-label ─── */
.sub-label {
    font-size: 11px;
    font-weight: 600;
    letter-spacing: 0.9px;
    text-transform: uppercase;
    color: #6E7681;
    margin: 1.5rem 0 0.75rem;
}

/* ─── Item entry ─── */
.item-title {
    font-size: 13px;
    font-weight: 600;
    color: #C9D1D9;
    padding: 10px 0 6px;
    border-top: 1px solid #21262D;
    margin-top: 8px;
}
.factor-note {
    font-size: 12px;
    color: #6E7681;
    padding-top: 30px;
}

/* ─── Inputs ─── */
.stTextInput input,
.stNumberInput input,
.stTextArea textarea,
[data-testid="stDateInput"] input,
[data-baseweb="input"] input {
    background-color: #161B22 !important;
    border: 1px solid #30363D !important;
    color: #E6EDF3 !important;
    border-radius: 6px !important;
    font-family: 'Inter', sans-serif !important;
    font-size: 14px !important;
    transition: border-color 150ms ease !important;
}
.stTextInput input:focus,
.stNumberInput input:focus,
.stTextArea textarea:focus {
    border-color: #BE1E2D !important;
    box-shadow: 0 0 0 3px rgba(190, 30, 45, 0.15) !important;
    outline: none !important;
}

/* Selectbox */
[data-baseweb="select"] > div,
[data-testid="stSelectbox"] [data-baseweb="select"] {
    background-color: #161B22 !important;
    border: 1px solid #30363D !important;
    border-radius: 6px !important;
    color: #E6EDF3 !important;
}

/* File uploader */
[data-testid="stFileUploader"] section {
    background-color: #161B22 !important;
    border: 1px dashed #30363D !important;
    border-radius: 8px !important;
    transition: border-color 200ms ease, background-color 200ms ease !important;
}
[data-testid="stFileUploader"] section:hover {
    border-color: #BE1E2D !important;
    background-color: #1C1012 !important;
}
[data-testid="stFileUploader"] section p,
[data-testid="stFileUploader"] section span {
    color: #8B949E !important;
}

/* Multiselect */
[data-baseweb="tag"] {
    background-color: #1F2937 !important;
    border: 1px solid #374151 !important;
    color: #E6EDF3 !important;
    border-radius: 4px !important;
}
[data-testid="stMultiSelect"] [data-baseweb="select"] > div {
    background-color: #161B22 !important;
    border: 1px solid #30363D !important;
    border-radius: 6px !important;
}

/* Checkbox */
[data-testid="stCheckbox"] label p {
    color: #8B949E !important;
    font-size: 13px !important;
}

/* Labels */
[data-testid="stWidgetLabel"] p,
.stTextInput label p,
.stNumberInput label p,
.stTextArea label p,
.stSelectbox label p,
.stMultiSelect label p,
.stCheckbox label p,
.stDateInput label p {
    color: #8B949E !important;
    font-size: 13px !important;
    font-weight: 500 !important;
}

/* ─── Primary button ─── */
.stButton button[kind="primary"] {
    background-color: #BE1E2D !important;
    border: none !important;
    color: #ffffff !important;
    font-family: 'Lexend', sans-serif !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    border-radius: 6px !important;
    padding: 10px 28px !important;
    letter-spacing: 0.1px !important;
    transition: background-color 150ms ease, box-shadow 150ms ease !important;
    cursor: pointer !important;
}
.stButton button[kind="primary"]:hover {
    background-color: #9B1823 !important;
    box-shadow: 0 0 0 3px rgba(190, 30, 45, 0.25) !important;
}
.stButton button[kind="secondary"] {
    background-color: #21262D !important;
    border: 1px solid #30363D !important;
    color: #E6EDF3 !important;
    border-radius: 6px !important;
    font-family: 'Inter', sans-serif !important;
    transition: background-color 150ms ease !important;
}
.stButton button[kind="secondary"]:hover {
    background-color: #30363D !important;
}

/* ─── Download button ─── */
.stDownloadButton button {
    background-color: #BE1E2D !important;
    border: none !important;
    color: #ffffff !important;
    font-family: 'Lexend', sans-serif !important;
    font-weight: 600 !important;
    font-size: 15px !important;
    border-radius: 8px !important;
    padding: 13px 28px !important;
    width: 100% !important;
    transition: background-color 150ms ease, box-shadow 150ms ease !important;
    cursor: pointer !important;
}
.stDownloadButton button:hover {
    background-color: #9B1823 !important;
    box-shadow: 0 0 0 3px rgba(190, 30, 45, 0.25) !important;
}

/* ─── Expanders ─── */
[data-testid="stExpander"] {
    background-color: #161B22 !important;
    border: 1px solid #30363D !important;
    border-radius: 8px !important;
    margin-bottom: 8px !important;
}
[data-testid="stExpander"] summary {
    font-family: 'Lexend', sans-serif !important;
    font-size: 14px !important;
    font-weight: 600 !important;
    color: #C9D1D9 !important;
    padding: 14px 18px !important;
}
[data-testid="stExpander"] summary:hover {
    color: #E6EDF3 !important;
    background-color: #21262D !important;
    border-radius: 8px !important;
}

/* ─── Dataframe ─── */
[data-testid="stDataFrame"] {
    border: 1px solid #30363D !important;
    border-radius: 8px !important;
    overflow: hidden !important;
}
[data-testid="stDataFrame"] th {
    background-color: #161B22 !important;
    color: #8B949E !important;
    font-size: 12px !important;
    font-weight: 600 !important;
    letter-spacing: 0.5px !important;
    text-transform: uppercase !important;
}
[data-testid="stDataFrame"] td {
    color: #C9D1D9 !important;
    font-size: 13px !important;
}

/* ─── Metric ─── */
[data-testid="stMetric"] {
    background-color: #161B22 !important;
    border: 1px solid #30363D !important;
    border-radius: 8px !important;
    padding: 14px 16px !important;
}
[data-testid="stMetricValue"] {
    color: #BE1E2D !important;
    font-family: 'Lexend', sans-serif !important;
    font-weight: 700 !important;
    font-size: 22px !important;
}
[data-testid="stMetricLabel"] {
    color: #6E7681 !important;
    font-size: 11px !important;
    font-weight: 600 !important;
    text-transform: uppercase !important;
    letter-spacing: 0.7px !important;
}

/* ─── Divider ─── */
hr {
    border: none !important;
    border-top: 1px solid #21262D !important;
    margin: 2rem 0 !important;
}

/* ─── Scrollbar ─── */
::-webkit-scrollbar { width: 6px; height: 6px; }
::-webkit-scrollbar-track { background: #0D1117; }
::-webkit-scrollbar-thumb { background: #30363D; border-radius: 3px; }
::-webkit-scrollbar-thumb:hover { background: #484F58; }

/* ─── Spinner ─── */
[data-testid="stSpinner"] { color: #BE1E2D !important; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown(f"""
<div class="pcc-header">
    <div class="h-logos">
        <img src="{LOGO_PCC}" alt="PCC" style="height:52px;object-fit:contain;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.6));">
    </div>
    <div class="h-title">
        <h1>Informe Diario de Proyectos</h1>
        <p>Ecopetrol &nbsp;&middot;&nbsp; Contrato CW309754 &nbsp;&middot;&nbsp; Protección Catódica de Colombia</p>
    </div>
    <div class="h-logos">
        <img src="{LOGO_ECO}" alt="Ecopetrol" style="height:36px;object-fit:contain;filter:drop-shadow(0 2px 8px rgba(0,0,0,0.6));">
    </div>
</div>
""", unsafe_allow_html=True)

# ── Session state ─────────────────────────────────────────────────────────────
for key, default in [
    ("ff_data", None), ("report_wb", None),
    ("report_bytes", None), ("report_items", []), ("ff_photos", []),
]:
    if key not in st.session_state:
        st.session_state[key] = default

# ─────────────────────────────────────────────────────────────────────────────
# PASO 1 — Cargar archivos
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="step-hdr">
    <div><span class="step-num">01</span></div>
    <div>
        <div class="step-title">Cargar archivos</div>
        <div class="step-desc">Submission de FastField y plantilla del último reporte enviado</div>
    </div>
</div>
""", unsafe_allow_html=True)

col_a, col_b = st.columns(2, gap="large")

with col_a:
    st.markdown('<div class="upload-label">Submission FastField</div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-desc">Archivo exportado desde FastField (.xlsx)</div>', unsafe_allow_html=True)
    ff_file = st.file_uploader(
        "Submission", type=["xlsx"], key="ff_upload", label_visibility="collapsed",
    )
    if ff_file:
        try:
            ff_bytes = ff_file.read()
            st.session_state.ff_data = parse_submission(io.BytesIO(ff_bytes))
            fd = st.session_state.ff_data

            # Descargar fotos desde FastField API si hay credenciales configuradas
            photo_filenames = [p["filename"] for p in fd.get("fotos", [])]
            ff_email    = st.secrets.get("fastfield_email", "")
            ff_password = st.secrets.get("fastfield_password", "")
            ff_org_id   = st.secrets.get("fastfield_org_id", "")
            ff_sub_key  = st.secrets.get("fastfield_subscription_key", "")
            if photo_filenames and ff_email and ff_password:
                with st.spinner(f"Descargando {len(photo_filenames)} foto(s) desde FastField..."):
                    photo_bytes_list, api_err = download_submission_photos(
                        photo_filenames, ff_email, ff_password, ff_org_id, ff_sub_key
                    )
                st.session_state.ff_photos = photo_bytes_list
                n_ok = sum(1 for b in photo_bytes_list if b)
                if n_ok > 0:
                    st.markdown(
                        f'<div class="pill pill-ok"><span class="pill-dot"></span>'
                        f'Cargado &mdash; <strong>{fd.get("fecha_informe")}</strong> &nbsp;·&nbsp; '
                        f'{fd.get("locacion")} &nbsp;·&nbsp; {n_ok}/{len(photo_filenames)} fotos descargadas</div>',
                        unsafe_allow_html=True,
                    )
                else:
                    st.markdown(
                        f'<div class="pill pill-warn"><span class="pill-dot"></span>'
                        f'Cargado &mdash; fotos no descargadas. Error: <code>{api_err}</code></div>',
                        unsafe_allow_html=True,
                    )
            else:
                st.session_state.ff_photos = []
                st.markdown(
                    f'<div class="pill pill-ok"><span class="pill-dot"></span>'
                    f'Cargado &mdash; <strong>{fd.get("fecha_informe")}</strong> &nbsp;·&nbsp; {fd.get("locacion")}</div>',
                    unsafe_allow_html=True,
                )
        except Exception as e:
            st.markdown(
                f'<div class="pill pill-err"><span class="pill-dot"></span>Error: {e}</div>',
                unsafe_allow_html=True,
            )

with col_b:
    st.markdown('<div class="upload-label">Reporte formal (plantilla)</div>', unsafe_allow_html=True)
    st.markdown('<div class="upload-desc">Último reporte enviado a Ecopetrol (.xlsx)</div>', unsafe_allow_html=True)
    rpt_file = st.file_uploader(
        "Plantilla", type=["xlsx"], key="rpt_upload", label_visibility="collapsed",
    )
    if rpt_file:
        try:
            rpt_bytes = rpt_file.read()
            st.session_state.report_bytes = rpt_bytes
            st.session_state.report_wb = openpyxl.load_workbook(
                io.BytesIO(rpt_bytes), data_only=True
            )
            st.session_state.report_items = read_c_control_items(st.session_state.report_wb)
            current_no = read_reporte_no(st.session_state.report_wb)
            st.markdown(
                f'<div class="pill pill-ok"><span class="pill-dot"></span>'
                f'Reporte N.° <strong>{current_no}</strong> &nbsp;·&nbsp; '
                f'<strong>{len(st.session_state.report_items)}</strong> ítems en C.Control</div>',
                unsafe_allow_html=True,
            )
        except Exception as e:
            st.markdown(
                f'<div class="pill pill-err"><span class="pill-dot"></span>Error: {e}</div>',
                unsafe_allow_html=True,
            )

if st.session_state.ff_data is None or st.session_state.report_wb is None:
    st.markdown(
        '<div class="pill pill-warn" style="margin-top:12px;"><span class="pill-dot"></span>'
        'Carga ambos archivos para continuar.</div>',
        unsafe_allow_html=True,
    )
    st.stop()

ff = st.session_state.ff_data
wb = st.session_state.report_wb
current_reporte_no = read_reporte_no(wb)

# ─────────────────────────────────────────────────────────────────────────────
# PASO 2 — Datos generales
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="step-hdr">
    <div><span class="step-num">02</span></div>
    <div>
        <div class="step-title">Datos generales</div>
        <div class="step-desc">Encabezado del informe, narrativas y personal en campo</div>
    </div>
</div>
""", unsafe_allow_html=True)

c1, c2, c3 = st.columns(3, gap="medium")
with c1:
    fecha_informe = st.date_input("Fecha del informe", value=ff["fecha_informe"] or date.today())
with c2:
    reporte_no = st.number_input("Número de reporte", min_value=1, value=current_reporte_no + 1, step=1)
with c3:
    contrato = st.text_input("Contrato / OS", value=ff.get("contrato", "CW309754"))

c4, c5 = st.columns(2, gap="medium")
with c4:
    profesional = st.text_input("Profesional líder PCC", value=ff.get("profesional_lider", ""))
with c5:
    locacion_display = st.text_input("Locación", value=ff.get("locacion", "Cusiana"))

charla_diaria = st.text_input("Charla diaria", value=ff.get("charla_diaria", ""))

st.markdown('<div class="sub-label">Narrativas de avance</div>', unsafe_allow_html=True)
loc_label = ff.get("locacion", "Locación")

# Texto HSE con plantilla fija — solo cambia la charla del día
_charla = ff.get("charla_diaria", "").strip()
_charla_linea = f"Charla pre-operacional: {_charla}" if _charla else "Charla pre-operacional:"
_hse_default = (
    "Aseguramiento del área: Inspección y adecuación de las condiciones de seguridad en el entorno de trabajo.\n\n"
    f"{_charla_linea}\n\n"
    "Gestión HSE: Actualización y cumplimiento de los lineamientos establecidos en el plan de Salud, Seguridad y Ambiente.\n\n"
    "Trámites administrativos: Gestión y validación de permisos de trabajo para el inicio de actividades."
)

nc1, nc2 = st.columns(2, gap="medium")
with nc1:
    avance_cusiana = st.text_area(
        f"Avance relevante — {loc_label}", value=ff.get("avance_items_texto", ""), height=160,
    )
    avance_hse = st.text_area(
        f"Actividades HSE — {loc_label}", value=_hse_default, height=180,
    )
with nc2:
    administrativo = st.text_area(
        f"Administrativo — {loc_label}", value=ff.get("administrativo", ""), height=120,
    )

st.markdown('<div class="sub-label">Personal en campo y recursos</div>', unsafe_allow_html=True)
pc1, pc2, pc3, pc4 = st.columns(4, gap="medium")
with pc1:
    cal_region    = st.number_input("Calificado — Región",       min_value=0, value=1, step=1)
    no_cal_region = st.number_input("No calificado — Región",    min_value=0, value=2, step=1)
with pc2:
    cal_no_region    = st.number_input("Calificado — Fuera región",    min_value=0, value=7, step=1)
    no_cal_no_region = st.number_input("No calificado — Fuera región", min_value=0, value=0, step=1)
with pc3:
    maquinaria = st.number_input("Maquinaria (und.)", min_value=0, value=1, step=1)
with pc4:
    equipo = st.number_input("Equipo (und.)", min_value=0, value=1, step=1)

hh_dia = st.number_input(
    "HH registradas en el día",
    min_value=0.0,
    value=float(ff.get("horas_hombre") or 0.0),
    step=0.5,
)

# ─────────────────────────────────────────────────────────────────────────────
# PASO 3 — HSE
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="step-hdr">
    <div><span class="step-num">03</span></div>
    <div>
        <div class="step-title">Registro fotográfico</div>
        <div class="step-desc">Hasta 6 fotos — solo edita la descripción de cada una</div>
    </div>
</div>
""", unsafe_allow_html=True)

st.markdown(
    '<p style="font-size:13px;color:#6E7681;margin:0 0 16px;">Carga las fotos descargadas de FastField. '
    'La fecha y ubicación se llenan automáticamente.</p>',
    unsafe_allow_html=True,
)

fotos_data = []
FOTO_SLOTS  = 6
ff_photos   = st.session_state.ff_photos   # fotos descargadas automáticamente

# Fotos disponibles descargadas, con su número original (1-based). Se descartan
# las que fallaron al descargar (None).
available   = [(i + 1, b) for i, b in enumerate(ff_photos) if b]
n_avail     = len(available)
photo_by_num = {num: b for num, b in available}
NONE_LABEL  = "— Sin foto / cargar manual —"
opt_labels  = [NONE_LABEL] + [f"Foto {num}" for num, _ in available]

if n_avail:
    extra = " El formulario trae más de 6 fotos: elige cuáles 6 van al Excel." if n_avail > FOTO_SLOTS else ""
    st.markdown(
        f'<div class="pill pill-ok" style="margin-bottom:12px;"><span class="pill-dot"></span>'
        f'{n_avail} foto(s) descargada(s) desde FastField.{extra} '
        f'En cada espacio elige qué foto va y escribe su descripción.</div>',
        unsafe_allow_html=True,
    )
else:
    st.markdown(
        '<p style="font-size:13px;color:#6E7681;margin:0 0 16px;">Carga las fotos '
        'descargadas de FastField. La fecha y ubicación se llenan automáticamente.</p>',
        unsafe_allow_html=True,
    )

_chosen_nums = []  # para avisar si se repite la misma foto en dos slots

for row_idx in range(0, FOTO_SLOTS, 2):
    fc1, fc2 = st.columns(2, gap="large")
    for col_idx, fc in enumerate([fc1, fc2]):
        slot = row_idx + col_idx + 1
        if slot > FOTO_SLOTS:
            break
        with fc:
            st.markdown(f'<div class="sub-label">Foto {slot}</div>', unsafe_allow_html=True)

            img_bytes = None

            if n_avail:
                # Selector: por defecto el slot toma la foto número {slot}.
                default_label = f"Foto {slot}" if slot in photo_by_num else NONE_LABEL
                choice = st.selectbox(
                    f"Seleccionar foto para espacio {slot}",
                    opt_labels,
                    index=opt_labels.index(default_label),
                    key=f"sel_{slot}",
                    label_visibility="collapsed",
                )
                if choice != NONE_LABEL:
                    num = int(choice.split()[1])
                    img_bytes = photo_by_num.get(num)
                    _chosen_nums.append(num)
                    if img_bytes:
                        st.image(img_bytes, use_container_width=True)

            if img_bytes is None:
                # Sin selección (o sin fotos automáticas): permitir carga manual.
                img_file = st.file_uploader(
                    f"Foto {slot}",
                    type=["jpg", "jpeg", "png"],
                    key=f"foto_{slot}",
                    label_visibility="collapsed",
                )
                if img_file:
                    img_bytes = img_file.read()
                    st.image(img_bytes, use_container_width=True)
                else:
                    st.markdown(
                        '<div style="height:140px;background:#161B22;border:1px dashed #30363D;'
                        'border-radius:8px;display:flex;align-items:center;justify-content:center;'
                        'color:#484F58;font-size:12px;">Sin foto</div>',
                        unsafe_allow_html=True,
                    )

            desc = st.text_input(
                "Descripción",
                value="",
                placeholder="Ej: Caja de cabezal de Pozo CPR 003",
                key=f"desc_{slot}",
            )
            fotos_data.append({"image_bytes": img_bytes, "descripcion": desc})

# Aviso si la misma foto quedó elegida en más de un espacio.
_dups = {n for n in _chosen_nums if _chosen_nums.count(n) > 1}
if _dups:
    st.warning(
        "⚠️ La(s) Foto(s) " + ", ".join(f"#{n}" for n in sorted(_dups)) +
        " está(n) seleccionada(s) en más de un espacio. Revisa para no repetir fotos en el Excel."
    )

st.markdown("""
<div class="step-hdr">
    <div><span class="step-num">04</span></div>
    <div>
        <div class="step-title">Indicadores HSE</div>
        <div class="step-desc">Eventos de seguridad y salud registrados durante el día</div>
    </div>
</div>
""", unsafe_allow_html=True)

h1, h2, h3 = st.columns(3, gap="medium")
with h1:
    hse_accid_cpt  = st.number_input("Accidentes con pérdida de tiempo", min_value=0, value=0)
    hse_primeros   = st.number_input("Primeros auxilios",                min_value=0, value=0)
    hse_incid_vial = st.number_input("Incidentes viales",                min_value=0, value=0)
with h2:
    hse_accid_spt  = st.number_input("Accidentes sin pérdida de tiempo", min_value=0, value=0)
    hse_derrames   = st.number_input("Derrames",                         min_value=0, value=0)
    hse_casi_ac    = st.number_input("Casi accidentes",                  min_value=0, value=0)
with h3:
    hse_fallas     = st.number_input("Fallas de control",                min_value=0, value=0)
    hse_aseguramto = st.number_input("Aseguramiento de comportamientos", min_value=0, value=0)
    hse_visitas    = st.number_input("Visitas gerenciales",              min_value=0, value=0)

# ─────────────────────────────────────────────────────────────────────────────
# PASO 5 — Avance por ítems
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="step-hdr">
    <div><span class="step-num">05</span></div>
    <div>
        <div class="step-title">Avance por ítems</div>
        <div class="step-desc">Cantidades ejecutadas hoy para actualizar la Curva S en C.Control</div>
    </div>
</div>
""", unsafe_allow_html=True)

ff_locacion    = ff.get("locacion", "")
active_section = detect_section(ff_locacion)

wb_data      = openpyxl.load_workbook(io.BytesIO(st.session_state.report_bytes), data_only=True)
date_col_num = find_date_column(wb_data, fecha_informe)

ic1, ic2 = st.columns(2, gap="medium")
with ic1:
    if active_section:
        st.markdown(
            f'<div class="pill pill-ok"><span class="pill-dot"></span>'
            f'Locación: <strong>{SECTION_LABELS[active_section]}</strong> &nbsp;·&nbsp; sección <strong>{active_section}</strong> en C.Control</div>',
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            f'<div class="pill pill-warn"><span class="pill-dot"></span>'
            f'Locación "{ff_locacion}" no reconocida. Selecciona manualmente.</div>',
            unsafe_allow_html=True,
        )
        active_section = st.selectbox(
            "Locación", options=list(SECTION_LABELS.keys()),
            format_func=lambda k: SECTION_LABELS[k],
        )
with ic2:
    if date_col_num is None:
        st.markdown(
            f'<div class="pill pill-err"><span class="pill-dot"></span>'
            f'Fecha {fecha_informe} sin columna en C.Control.</div>',
            unsafe_allow_html=True,
        )
    else:
        from openpyxl.utils import get_column_letter
        st.markdown(
            f'<div class="pill pill-info"><span class="pill-dot"></span>'
            f'Fecha <strong>{fecha_informe}</strong> &nbsp;·&nbsp; columna <strong>{get_column_letter(date_col_num)}</strong> en C.Control</div>',
            unsafe_allow_html=True,
        )

if ff.get("avance_items_texto"):
    with st.expander("Notas de campo (referencia)", expanded=False):
        st.text(ff["avance_items_texto"])

# ── Ítems de FastField (subform_4) ────────────────────────────────────────────
ff_items     = ff.get("items_fastfield", [])          # lista de {item_num, cantidad, unidad}
ff_items_map = {it["item_num"]: it for it in ff_items} # {23: {...}, 8: {...}}

items_all     = st.session_state.report_items
sections      = ["CUSIANA", "FLORENA", "CUPIAGUA"]
all_item_qtys = []

# ── Mostrar resumen auto-detectado si hay ítems en subform_4 ─────────────────
if ff_items:
    st.markdown(
        f'<div class="pill pill-ok" style="margin-bottom:12px;"><span class="pill-dot"></span>'
        f'FastField reportó <strong>{len(ff_items)}</strong> ítem(s) con cantidades — se pre-cargan automáticamente.</div>',
        unsafe_allow_html=True,
    )
    df_ff = pd.DataFrame([{
        "Ítem N°":     it["item_num"],
        "Descripción": it["descripcion"][:70],
        "Cantidad":    it["cantidad"],
        "Unidad":      it["unidad"],
    } for it in ff_items])
    st.dataframe(df_ff, use_container_width=True, hide_index=True)
else:
    st.markdown(
        '<p style="font-size:13px;color:#6E7681;margin:8px 0 16px;line-height:1.6;">'
        'Selecciona los ítems ejecutados hoy e ingresa la cantidad. '
        'Para ítems en M3 con reporte en metros lineales, ajusta el factor (ML &times; factor = M3 contrato).</p>',
        unsafe_allow_html=True,
    )

for section in sections:
    section_items = [it for it in items_all if it["section"] == section]
    if not section_items:
        continue

    is_active = section == active_section
    label     = SECTION_LABELS.get(section, section)

    # Ítems de esta sección que vienen pre-cargados desde FastField
    section_ff_nums = {
        it["item_num"] for it in ff_items
        if any(ci["item"] == it["item_num"] and ci["section"] == section for ci in section_items)
    }
    badge = f"  (locación del día)" if is_active else ""

    with st.expander(f"{label}{badge}  —  {len(section_items)} ítems", expanded=is_active):
        rows = []
        for it in section_items:
            acum     = it.get("acumulado") or 0
            has_prev = isinstance(acum, (int, float)) and acum > 0
            ff_qty   = ff_items_map.get(it["item"]) if (is_active and it["item"] in ff_items_map) else None
            rows.append({
                "FF":          "Auto" if ff_qty else ("Prev" if has_prev else ""),
                "N°":          it["item"],
                "Esp.":        it["especialidad"],
                "Descripción": it["descripcion"][:80],
                "Und.":        it["unidad"],
                "Cant. total": it["cantidad_total"],
                "Acumulado":   acum,
                "_row_num":    it["row_num"],
                "_unidad":     it["unidad"],
                "_descripcion":it["descripcion"],

            })

        df   = pd.DataFrame(rows)
        disp = ["FF", "N°", "Esp.", "Descripción", "Und.", "Cant. total", "Acumulado"]

        # FastField solo aplica a la sección de la locación del día
        ff_nums_in_section = (
            [it["item"] for it in section_items if it["item"] in ff_items_map]
            if is_active else []
        )

        solo_prev = st.checkbox("Solo ítems con avance acumulado previo", key=f"chk_{section}")
        df_show   = df[df["FF"].isin(["Auto", "Prev"])] if solo_prev else df

        if df_show.empty:
            st.markdown('<p style="color:#6E7681;font-size:13px;padding:8px 0;">Sin ítems en este filtro.</p>', unsafe_allow_html=True)
            continue

        st.dataframe(df_show[disp], use_container_width=True, hide_index=True)

        # Pre-selección solo en la sección activa
        default_sel = [n for n in ff_nums_in_section if n in df_show["N°"].tolist()]
        selected = st.multiselect(
            "Ítems ejecutados hoy",
            options=df_show["N°"].tolist(),
            default=default_sel,
            key=f"sel_{section}",
            format_func=lambda n, _df=df_show: f"#{n}  {_df[_df['N°']==n]['Descripción'].values[0]}",
        )

        for item_no in selected:
            item_row = df[df["N°"] == item_no].iloc[0]
            unidad   = item_row["_unidad"]
            desc     = item_row["_descripcion"]
            row_num  = item_row["_row_num"]

            # Cantidad pre-cargada desde FastField solo si es la sección activa
            ff_entry    = ff_items_map.get(item_no) if is_active else None
            qty_default = float(ff_entry["cantidad"]) if ff_entry else 0.0
            auto_label  = " (FastField)" if ff_entry else ""

            st.markdown(
                f'<div class="item-title">#{item_no} &nbsp;—&nbsp; {desc[:100]}</div>',
                unsafe_allow_html=True,
            )
            qc1, qc2 = st.columns([2, 2], gap="medium")
            with qc1:
                qty = st.number_input(
                    f"Cantidad{auto_label} ({unidad})",
                    min_value=0.0, value=qty_default, step=0.001,
                    format="%.3f",
                    key=f"qty_{section}_{item_no}",
                )
            with qc2:
                st.metric(f"A registrar ({unidad})", f"{qty:.3f}")

            all_item_qtys.append({
                "section":        section,
                "item_no":        item_no,
                "row_num":        int(row_num),
                "cantidad_campo": qty,
                "factor":         1.0,
                "unidad":         unidad,
                "cantidad_final": round(qty, 4),
            })

# ─────────────────────────────────────────────────────────────────────────────
# PASO 5 — Generar informe
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="step-hdr">
    <div><span class="step-num">06</span></div>
    <div>
        <div class="step-title">Generar informe</div>
        <div class="step-desc">Revisa el resumen y descarga el Excel listo para enviar a Ecopetrol</div>
    </div>
</div>
""", unsafe_allow_html=True)

items_con_qty = [q for q in all_item_qtys if q["cantidad_campo"] > 0]

if items_con_qty:
    st.markdown('<div class="sub-label">Cantidades a registrar en C.Control</div>', unsafe_allow_html=True)
    df_resumen = pd.DataFrame([{
        "Sección":            q["section"],
        "Ítem":               q["item_no"],
        "Cant. campo":        q["cantidad_campo"],
        "Factor":             q["factor"],
        "A registrar":        q["cantidad_final"],
        "Unidad contrato":    q["unidad"],
    } for q in items_con_qty])
    st.dataframe(df_resumen, use_container_width=True, hide_index=True)
else:
    st.markdown(
        '<p style="font-size:13px;color:#6E7681;margin-bottom:16px;">'
        'Sin cantidades de avance. El informe se generará solo con los datos generales.</p>',
        unsafe_allow_html=True,
    )

gen_col, _ = st.columns([2, 3])
with gen_col:
    generar = st.button("Generar informe", type="primary", use_container_width=True)

if generar:
    # HH total del día = total personas × horas por persona
    _total_personas = int(cal_region) + int(cal_no_region) + int(no_cal_region) + int(no_cal_no_region)
    _hh_total = _total_personas * hh_dia

    # ── Avance real acumulado ──────────────────────────────────────────────────
    # Formula: row9(date_col) = sum(qty*precio)/I13 ; row8 = row9 + row8_prev
    _avance_acum = None
    _avance_debug = []
    try:
        _ws_cc = wb["C.Control"]
        _I13 = _ws_cc.cell(row=13, column=9).value  # I13 = valor total contrato
        if _I13 and float(_I13) > 0:
            _price_map = {
                item["row_num"]: float(item["valor_unitario"] or 0)
                for item in st.session_state.report_items
            }
            _today_value = sum(
                q["cantidad_final"] * _price_map.get(q["row_num"], 0)
                for q in items_con_qty
            )
            _today_pct = _today_value / float(_I13)
            _avance_debug.append(f"Valor hoy: {_today_value:,.2f} / {float(_I13):,.2f} = {_today_pct:.6f}")

            _prev_acum = 0.0
            _date_col_prev = find_date_column(wb, fecha_informe)
            if _date_col_prev:
                _prev_col = _date_col_prev - 1
                if _prev_col >= 16:
                    _v = _ws_cc.cell(row=8, column=_prev_col).value
                    _prev_acum = float(_v) if isinstance(_v, (int, float)) else 0.0
                    _avance_debug.append(f"Acum. anterior (col {_prev_col}): {_prev_acum:.6f} (raw: {_v})")
            else:
                for _c in range(_ws_cc.max_column, 15, -1):
                    _v = _ws_cc.cell(row=8, column=_c).value
                    if isinstance(_v, (int, float)) and _v > 0:
                        _prev_acum = float(_v)
                        _avance_debug.append(f"Acum. anterior (col {_c}): {_prev_acum:.6f}")
                        break

            _avance_acum = _prev_acum + _today_pct
            _avance_debug.append(f"Total: {_prev_acum:.6f} + {_today_pct:.6f} = {_avance_acum:.6f} ({_avance_acum*100:.2f}%)")
        else:
            _avance_debug.append(f"I13 = {_I13} — sin valor de contrato para calcular avance")
    except Exception as _avance_err:
        _avance_debug.append(f"Error: {_avance_err}")

    form_data = {
        "fecha_informe":                    fecha_informe,
        "reporte_no":                       int(reporte_no),
        "charla_diaria":                    charla_diaria,
        "avance_cusiana":                   avance_cusiana,
        "administrativo_cusiana":           administrativo,
        "avance_hse_cusiana":               avance_hse,
        "personal_calificado_region":       int(cal_region),
        "personal_calificado_no_region":    int(cal_no_region),
        "personal_no_calificado_region":    int(no_cal_region),
        "personal_no_calificado_no_region": int(no_cal_no_region),
        "maquinaria":                       int(maquinaria),
        "equipo":                           int(equipo),
        "hh_dia":                           _hh_total,   # Total HH del día para la hoja HSE
        "hse_accid_cpt":                    int(hse_accid_cpt),
        "hse_accid_spt":                    int(hse_accid_spt),
        "hse_primeros_aux":                 int(hse_primeros),
        "hse_derrames":                     int(hse_derrames),
        "hse_incid_viales":                 int(hse_incid_vial),
        "hse_casi_accid":                   int(hse_casi_ac),
        "hse_fallas_ctrl":                  int(hse_fallas),
        "hse_aseguramiento":                int(hse_aseguramto),
        "hse_visitas_ger":                  int(hse_visitas),
        "fotos":                            fotos_data,
        "locacion_display":                 locacion_display,
        "avance_real_acumulado":            _avance_acum,   # decimal, ej. 0.1828
    }

    with st.spinner("Generando informe..."):
        try:
            output_bytes = update_report(
                io.BytesIO(st.session_state.report_bytes),
                form_data,
                items_con_qty,
            )
            fecha_str = fecha_informe.strftime("%d-%m-%Y")
            filename  = f"{int(reporte_no)}_{contrato}_Reporte_Diario_{fecha_str}.xlsx"

            st.markdown(
                '<div class="pill pill-ok" style="margin-bottom:14px;">'
                '<span class="pill-dot"></span>'
                'Informe generado correctamente. Descarga el archivo a continuación.</div>',
                unsafe_allow_html=True,
            )
            if _avance_debug:
                with st.expander("Detalle avance acumulado (debug)", expanded=False):
                    for line in _avance_debug:
                        st.text(line)
            dl_col, _ = st.columns([2, 3])
            with dl_col:
                st.download_button(
                    label="Descargar informe Excel",
                    data=output_bytes,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
        except Exception as e:
            st.markdown(
                f'<div class="pill pill-err"><span class="pill-dot"></span>Error al generar: {e}</div>',
                unsafe_allow_html=True,
            )
            st.exception(e)
